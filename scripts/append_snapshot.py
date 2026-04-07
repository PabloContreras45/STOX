#!/usr/bin/env python3
"""
append_snapshot.py — called by the Vite server endpoint /api/snapshots
Reads the snapshot rows from stdin as JSON, appends them to Tracking/tracking.xlsx
Usage: python3 append_snapshot.py < snapshot_rows.json
"""
import sys, json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell

TRACKING = os.path.join(os.path.dirname(__file__), '..', 'Tracking', 'tracking.xlsx')

BORDER_CLR = "E4E4E0"
PAPER      = "FFFFFF"
BG         = "FAFAF8"
INK        = "0A0A0A"
MUTED      = "6B6B6B"
FAINT      = "B0B0B0"

def thin():
    s = Side(style='thin', color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def write_cell(cell, val, fmt=None, bold=False, color=INK, align='left', bg=PAPER):
    cell.value = val
    cell.font  = Font(name='Inter', size=9, bold=bold, color=color)
    cell.fill  = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = thin()
    if fmt:
        cell.number_format = fmt

def main():
    raw = sys.stdin.read().strip()
    if not raw:
        print(json.dumps({"ok": False, "error": "No input"}))
        sys.exit(1)

    rows = json.loads(raw)  # list of snapshot objects

    wb = openpyxl.load_workbook(TRACKING)
    ws = wb["Snapshots"]

    # Unmerge any merged regions that start at row 3 or later (sample/note rows)
    to_unmerge = [str(rng) for rng in ws.merged_cells.ranges
                  if rng.min_row >= 3]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)

    # Find first truly empty row (skip MergedCell instances and non-empty cells)
    next_row = None
    for r in range(3, ws.max_row + 2):
        cell = ws.cell(row=r, column=2)  # ticker col
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if val is None or val == "" or val == "—":
            # Also check col 1 isn't a note row
            c1 = ws.cell(row=r, column=1)
            if isinstance(c1, MergedCell) or "↑" in str(c1.value or ""):
                continue
            next_row = r
            break
    if next_row is None:
        next_row = ws.max_row + 1

    added = 0
    for row_data in rows:
        bg_row = PAPER if (next_row % 2) == 1 else BG

        # Col 1: fecha_snapshot
        write_cell(ws.cell(next_row, 1), row_data.get("fecha_snapshot", ""), color=MUTED, bg=bg_row)
        # Col 2: ticker
        write_cell(ws.cell(next_row, 2), row_data.get("ticker", ""), bold=True, color=INK, bg=bg_row)
        # Col 3: bucket
        write_cell(ws.cell(next_row, 3), row_data.get("bucket", ""), color=MUTED, bg=bg_row)
        # Col 4: potencial_pct (store as decimal for % format)
        pot = row_data.get("potencial_pct")
        if pot is not None:
            write_cell(ws.cell(next_row, 4), pot / 100, fmt='0.0%', color=INK, bg=bg_row, align='right')
        else:
            write_cell(ws.cell(next_row, 4), None, color=FAINT, bg=bg_row, align='right')
            ws.cell(next_row, 4).value = "—"
        # Col 5: R/R
        rr = row_data.get("rr")
        if rr is not None:
            write_cell(ws.cell(next_row, 5), rr, fmt='0.00"x"', color=INK, bg=bg_row, align='right')
        else:
            write_cell(ws.cell(next_row, 5), "—", color=FAINT, bg=bg_row, align='right')
        # Col 6: subida_max_pct
        sub = row_data.get("subida_max_pct")
        if sub is not None:
            write_cell(ws.cell(next_row, 6), sub / 100, fmt='0.0%', color=INK, bg=bg_row, align='right')
        else:
            write_cell(ws.cell(next_row, 6), "—", color=FAINT, bg=bg_row, align='right')
        # Col 7: bajada_max_pct
        baj = row_data.get("bajada_max_pct")
        if baj is not None:
            write_cell(ws.cell(next_row, 7), baj / 100, fmt='0.0%', color=INK, bg=bg_row, align='right')
        else:
            write_cell(ws.cell(next_row, 7), "—", color=FAINT, bg=bg_row, align='right')
        # Col 8: precio_ref
        precio = row_data.get("precio_ref")
        if precio is not None:
            write_cell(ws.cell(next_row, 8), precio, fmt='€#,##0.00', color=INK, bg=bg_row, align='right')
        else:
            write_cell(ws.cell(next_row, 8), "—", color=FAINT, bg=bg_row, align='right')
        # Cols 9-12: future API fields — leave as —
        for c in range(9, 13):
            write_cell(ws.cell(next_row, c), "—", color=FAINT, bg="F5F5F5", align='right')

        ws.row_dimensions[next_row].height = 20
        next_row += 1
        added += 1

    wb.save(TRACKING)
    print(json.dumps({"ok": True, "added": added, "next_row": next_row}))

if __name__ == "__main__":
    main()
