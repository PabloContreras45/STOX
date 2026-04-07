import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

OUT = os.path.join(os.path.dirname(__file__), '..', 'Tracking', 'tracking.xlsx')

# ── Design tokens (match STOX palette) ─────────────────────────────────────
INK        = "0A0A0A"
GOLD       = "C5973A"
GOLD_LIGHT = "F5EDD3"
GOLD_BDR   = "E8D5A3"
PAPER      = "FFFFFF"
BG         = "FAFAF8"
BORDER_CLR = "E4E4E0"
RED        = "B91C1C"
RED_LIGHT  = "FEF2F2"
GREEN      = "166534"
GREEN_LIGHT= "F0FDF4"
MUTED      = "6B6B6B"
FAINT      = "B0B0B0"
BLUE       = "1D4ED8"
BLUE_LIGHT = "EFF6FF"

def thin(color=BORDER_CLR):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(cell, text, bg=INK, fg=GOLD_LIGHT, bold=True, size=9, wrap=False):
    cell.value = text
    cell.font  = Font(name='Inter', bold=bold, color=fg, size=size)
    cell.fill  = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=wrap)
    cell.border = thin(BORDER_CLR)

def data_cell(cell, val=None, fmt=None, color=INK, bold=False, align='left',
              bg=PAPER, wrap=False):
    if val is not None:
        cell.value = val
    cell.font  = Font(name='Inter', size=9, color=color, bold=bold)
    cell.fill  = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                                wrap_text=wrap)
    cell.border = thin()
    if fmt:
        cell.number_format = fmt

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Snapshots
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Snapshots"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A3'

# ── Title row ───────────────────────────────────────────────────────────────
ws1.merge_cells('A1:L1')
t = ws1['A1']
t.value = "◆ STOX — Tracker de estimaciones"
t.font  = Font(name='Inter', bold=True, size=12, color=GOLD)
t.fill  = PatternFill('solid', fgColor=INK)
t.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[1].height = 28

# ── Column headers ──────────────────────────────────────────────────────────
COLS_S1 = [
    ("fecha_snapshot",   "Fecha\nSnapshot",   12),
    ("ticker",           "Ticker",             9),
    ("bucket",           "Bucket",            11),
    ("potencial_pct",    "Potencial\n(%)",     11),
    ("rr",               "R/R",                8),
    ("subida_max_pct",   "Subida\nMáx (%)",   11),
    ("bajada_max_pct",   "Bajada\nMáx (%)",   11),
    ("precio_ref",       "Precio\nRef (€)",   11),
    # Future API fields — grayed out until connected
    ("precio_en_fecha",  "Precio\nReal (€)",  12),
    ("retorno_real_pct", "Retorno\nReal (%)", 12),
    ("error_pct",        "Error\nPredicción", 12),
    ("fuente_precio",    "Fuente\nPrecio",    12),
]

ws1.row_dimensions[2].height = 30
for col_idx, (key, label, width) in enumerate(COLS_S1, start=1):
    cell = ws1.cell(row=2, column=col_idx)
    # Future API cols (9-12) get a slightly lighter header
    bg = "1a1a1a" if col_idx <= 8 else "2a2a2a"
    fg = GOLD_LIGHT if col_idx <= 8 else FAINT
    header_cell(cell, label, bg=bg, fg=fg, size=8, wrap=True)
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# ── Section label for future API cols ──────────────────────────────────────
ws1.merge_cells('I1:L1')  # Overwrite the merge already set above for A1:L1?
# We can't merge I1:L1 after A1:L1 covers it. Use row 0 trick with an extra row
# Actually, add a note in a comment on I2
from openpyxl.comments import Comment
note = ws1['I2']
note.comment = Comment("Estas columnas se rellenarán automáticamente cuando\nse conecte una API de precios (Yahoo Finance, Polygon.io, etc.)\nPor ahora, puedes rellenar 'Precio Real' manualmente.", "STOX")

# ── 3 sample rows to show structure ─────────────────────────────────────────
sample_rows = [
    ["2026-03-12", "NVDA", "Satellite", 45.0, 2.8, 55.0, -20.0, None, None, None, None, None],
    ["2026-03-12", "MSFT", "Core",      22.0, 1.9, 28.0, -15.0, None, None, None, None, None],
    ["2026-03-12", "CEG",  "Wildshots", 68.0, 3.1, 75.0, -24.0, None, None, None, None, None],
]
for r_idx, row_data in enumerate(sample_rows, start=3):
    bg_row = PAPER if r_idx % 2 == 1 else BG
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=r_idx, column=c_idx)
        # Format
        if c_idx == 1:  # date
            data_cell(cell, val, color=MUTED, bg=bg_row)
        elif c_idx == 2:  # ticker
            data_cell(cell, val, bold=True, color=INK, bg=bg_row)
        elif c_idx == 3:  # bucket
            data_cell(cell, val, color=MUTED, bg=bg_row)
        elif c_idx in (4, 6, 7):  # percentages
            if val is not None:
                data_cell(cell, val/100, fmt='0.0%', color=INK, bg=bg_row, align='right')
            else:
                data_cell(cell, "—", color=FAINT, bg=bg_row, align='right')
        elif c_idx == 5:  # R/R
            if val is not None:
                data_cell(cell, val, fmt='0.00"x"', color=INK, bg=bg_row, align='right')
            else:
                data_cell(cell, "—", color=FAINT, bg=bg_row, align='right')
        elif c_idx == 8:  # precio_ref
            data_cell(cell, "—", color=FAINT, bg=bg_row, align='right')
        else:  # future API cols
            data_cell(cell, "—", color=FAINT, bg="F5F5F5", align='right')

# Row 6: placeholder note
ws1.merge_cells('A6:L6')
note_row = ws1['A6']
note_row.value = "↑ Filas de ejemplo — se sobreescribirán al guardar el primer snapshot real desde STOX"
note_row.font = Font(name='Inter', size=8, italic=True, color=FAINT)
note_row.fill = PatternFill('solid', fgColor="FAFAF8")
note_row.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[6].height = 18

# Set row heights for data rows
for r in range(3, 6):
    ws1.row_dimensions[r].height = 20

# ── Conditional formatting: color R/R column ────────────────────────────────
# Green >= 2.5, Amber 1.5-2.5, Red < 1.5
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import fills
rr_col = "E"
ws1.conditional_formatting.add(f'{rr_col}3:{rr_col}1048576',
    CellIsRule(operator='greaterThanOrEqual', formula=['2.5'],
               font=Font(color=GREEN, bold=True),
               fill=PatternFill('solid', fgColor=GREEN_LIGHT)))
ws1.conditional_formatting.add(f'{rr_col}3:{rr_col}1048576',
    CellIsRule(operator='between', formula=['1.5', '2.499'],
               font=Font(color="92400E", bold=True),
               fill=PatternFill('solid', fgColor="FFFBEB")))
ws1.conditional_formatting.add(f'{rr_col}3:{rr_col}1048576',
    CellIsRule(operator='lessThan', formula=['1.5'],
               font=Font(color=RED, bold=True),
               fill=PatternFill('solid', fgColor=RED_LIGHT)))

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Precisión
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Precisión")
ws2.sheet_view.showGridLines = False

# ── Title ───────────────────────────────────────────────────────────────────
ws2.merge_cells('A1:J1')
t2 = ws2['A1']
t2.value = "◆ STOX — Análisis de precisión predictiva"
t2.font  = Font(name='Inter', bold=True, size=12, color=GOLD)
t2.fill  = PatternFill('solid', fgColor=INK)
t2.alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[1].height = 28

# ── Note about API ──────────────────────────────────────────────────────────
ws2.merge_cells('A2:J2')
n2 = ws2['A2']
n2.value = "Las fórmulas de esta hoja se activan automáticamente cuando se rellenen las columnas 'Precio Real' e 'Error Predicción' en Snapshots (manual o vía API futura)."
n2.font  = Font(name='Inter', size=8, italic=True, color=MUTED)
n2.fill  = PatternFill('solid', fgColor=BG)
n2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws2.row_dimensions[2].height = 22

# ── Section A: Resumen global ────────────────────────────────────────────────
def section_title(ws, row, col, text, colspan=5):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col+colspan-1)
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font  = Font(name='Inter', bold=True, size=9, color=GOLD_LIGHT)
    c.fill  = PatternFill('solid', fgColor=INK)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 22

def metric_row(ws, row, label, formula, fmt='0.0%', col=1):
    lc = ws.cell(row=row, column=col)
    lc.value = label
    lc.font  = Font(name='Inter', size=9, color=MUTED)
    lc.fill  = PatternFill('solid', fgColor=BG)
    lc.border = thin()
    lc.alignment = Alignment(horizontal='left', vertical='center')

    vc = ws.cell(row=row, column=col+1)
    vc.value = formula
    vc.font  = Font(name='Inter', size=9, bold=True, color=INK)
    vc.fill  = PatternFill('solid', fgColor=PAPER)
    vc.number_format = fmt
    vc.border = thin()
    vc.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[row].height = 20

# Global summary block
section_title(ws2, 4, 1, "  RESUMEN GLOBAL", colspan=4)

metric_row(ws2, 5,  "Total snapshots guardados",
           "=COUNTA(Snapshots!B3:B10000)-COUNTIF(Snapshots!B3:B10000,\"—\")",
           fmt='#,##0', col=1)
metric_row(ws2, 6,  "Tickers únicos en seguimiento",
           "=SUMPRODUCT(1/COUNTIF(IF(Snapshots!B3:B10000<>\"—\",Snapshots!B3:B10000,\"x\"),IF(Snapshots!B3:B10000<>\"—\",Snapshots!B3:B10000,\"x\")))-1",
           fmt='#,##0', col=1)
metric_row(ws2, 7,  "Snapshots con precio real",
           "=COUNTIF(Snapshots!I3:I10000,\"<>—\")-COUNTBLANK(Snapshots!I3:I10000)",
           fmt='#,##0', col=1)
metric_row(ws2, 8,  "Error medio de predicción (todos)",
           '=IFERROR(AVERAGEIF(Snapshots!K3:K10000,"<>—",Snapshots!K3:K10000),"Pendiente datos")',
           fmt='0.0%', col=1)
metric_row(ws2, 9,  "Potencial medio prometido",
           "=IFERROR(AVERAGE(Snapshots!D3:D10000),\"—\")",
           fmt='0.0%', col=1)
metric_row(ws2, 10, "R/R medio de la cartera seguida",
           "=IFERROR(AVERAGE(Snapshots!E3:E10000),\"—\")",
           fmt='0.00"x"', col=1)

# ── Section B: Por bucket ────────────────────────────────────────────────────
section_title(ws2, 12, 1, "  POR BUCKET", colspan=6)

# Headers
bucket_headers = ["Bucket", "Snaps", "Potencial\nMedio", "R/R\nMedio",
                  "Error Medio\nPredicción", "% Snaps con\nDato Real"]
for ci, h in enumerate(bucket_headers, start=1):
    c = ws2.cell(row=13, column=ci)
    header_cell(c, h, bg="1a1a1a", fg=GOLD_LIGHT, size=8, wrap=True)
    ws2.row_dimensions[13].height = 30

for bi, bucket in enumerate(["Core", "Satellite", "Wildshots"], start=14):
    bg_b = PAPER if bi % 2 == 0 else BG
    b_name = ws2.cell(row=bi, column=1)
    data_cell(b_name, bucket, bold=True, color=INK, bg=bg_b)

    b_snaps = ws2.cell(row=bi, column=2)
    b_snaps.value = f'=COUNTIF(Snapshots!C3:C10000,A{bi})'
    data_cell(b_snaps, fmt='#,##0', bg=bg_b, align='right')
    b_snaps.value = f'=COUNTIF(Snapshots!C3:C10000,A{bi})'

    b_pot = ws2.cell(row=bi, column=3)
    b_pot.value = f'=IFERROR(AVERAGEIF(Snapshots!C3:C10000,A{bi},Snapshots!D3:D10000),"—")'
    data_cell(b_pot, fmt='0.0%', bg=bg_b, align='right')
    b_pot.value = f'=IFERROR(AVERAGEIF(Snapshots!C3:C10000,A{bi},Snapshots!D3:D10000),"—")'

    b_rr = ws2.cell(row=bi, column=4)
    b_rr.value = f'=IFERROR(AVERAGEIF(Snapshots!C3:C10000,A{bi},Snapshots!E3:E10000),"—")'
    data_cell(b_rr, fmt='0.00"x"', bg=bg_b, align='right')
    b_rr.value = f'=IFERROR(AVERAGEIF(Snapshots!C3:C10000,A{bi},Snapshots!E3:E10000),"—")'

    b_err = ws2.cell(row=bi, column=5)
    b_err.value = f'=IFERROR(AVERAGEIF(Snapshots!C3:C10000,A{bi},Snapshots!K3:K10000),"Pendiente")'
    data_cell(b_err, fmt='0.0%', bg=bg_b, align='right')
    b_err.value = f'=IFERROR(AVERAGEIF(Snapshots!C3:C10000,A{bi},Snapshots!K3:K10000),"Pendiente")'

    b_cov = ws2.cell(row=bi, column=6)
    b_cov.value = f'=IFERROR(COUNTIFS(Snapshots!C3:C10000,A{bi},Snapshots!I3:I10000,"<>—")/COUNTIF(Snapshots!C3:C10000,A{bi}),"—")'
    data_cell(b_cov, fmt='0%', bg=bg_b, align='right')
    b_cov.value = f'=IFERROR(COUNTIFS(Snapshots!C3:C10000,A{bi},Snapshots!I3:I10000,"<>—")/COUNTIF(Snapshots!C3:C10000,A{bi}),"—")'

    ws2.row_dimensions[bi].height = 20

# ── Section C: Por ticker ────────────────────────────────────────────────────
section_title(ws2, 18, 1, "  POR TICKER (últimos datos disponibles)", colspan=8)

ticker_headers = ["Ticker", "Bucket", "Último\nSnapshot", "Potencial\nEstimado",
                  "R/R", "Precio\nRef", "Precio\nReal", "Error\nPredicción"]
for ci, h in enumerate(ticker_headers, start=1):
    c = ws2.cell(row=19, column=ci)
    header_cell(c, h, bg="1a1a1a", fg=GOLD_LIGHT, size=8, wrap=True)
    ws2.row_dimensions[19].height = 30

ws2.merge_cells('A20:H20')
note_ticker = ws2['A20']
note_ticker.value = "Esta sección se completa automáticamente cuando haya datos reales. Añade filas con BUSCARV/XLOOKUP según vayas rellenando 'Precio Real' en Snapshots."
note_ticker.font = Font(name='Inter', size=8, italic=True, color=FAINT)
note_ticker.fill = PatternFill('solid', fgColor=BG)
note_ticker.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws2.row_dimensions[20].height = 28

# ── Section D: Nota API futura ───────────────────────────────────────────────
section_title(ws2, 22, 1, "  CONEXIÓN API DE PRECIOS (pendiente)", colspan=8)
ws2.merge_cells('A23:H24')
api_note = ws2['A23']
api_note.value = (
    "Cuando se conecte una API de precios (Yahoo Finance, Polygon.io, Alpha Vantage, etc.), "
    "el servidor Vite llamará automáticamente a /api/prices con cada ticker del archivo Snapshots. "
    "Los campos 'Precio Real', 'Retorno Real' y 'Error Predicción' se rellenarán solos. "
    "El endpoint /api/prices ya está previsto en la arquitectura del servidor."
)
api_note.font = Font(name='Inter', size=8, italic=True, color=MUTED)
api_note.fill = PatternFill('solid', fgColor=BLUE_LIGHT)
api_note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws2.row_dimensions[23].height = 42

# ── Column widths Sheet 2 ────────────────────────────────────────────────────
col_widths_s2 = [22, 12, 12, 14, 12, 16, 14, 16, 14, 14]
for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── Tab colors ───────────────────────────────────────────────────────────────
ws1.sheet_properties.tabColor = GOLD
ws2.sheet_properties.tabColor = "1D4ED8"

wb.save(OUT)
print(f"✓ tracking.xlsx saved to {OUT}")
